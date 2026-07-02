"""
Comparator — comparación de 2-4 perfiles estructurales.

Genera una tabla comparativa con:
- Filas: propiedades agrupadas por categoría (Identidad, Geometría,
  Eje fuerte, Eje débil, Torsión, Pandeo local, Materiales).
- Columnas: una por perfil, con la designación en el header.
- Resaltado: celdas donde el valor difiere del promedio de la fila
  en más de 5% (amarillo) o más de 20% (naranja).

Soporta exportación a:
- pandas DataFrame (to_dataframe)
- Excel .xlsx (to_excel, requiere openpyxl)
- Diccionario (to_dict)
- TSV al portapapeles (to_tsv)
"""
from __future__ import annotations

import math
from dataclasses import dataclass, field
from typing import Optional

from .models import Section


# ----------------------------------------------------------------------
# Estructura de la comparación
# ----------------------------------------------------------------------
# Cada entrada: (categoría, propiedad, label, getter, formatter)
# El getter es una función que recibe Section y retorna Optional[float] o str.
# El formatter es opcional; si es None, se usa formato automático.

def _get_float(name: str):
    """Retorna un getter que extrae un atributo float de Section."""
    def getter(s: Section) -> Optional[float]:
        v = getattr(s, name, None)
        return float(v) if v is not None else None
    return getter


def _get_str(name: str):
    """Retorna un getter que extrae un atributo str de Section."""
    def getter(s: Section) -> Optional[str]:
        v = getattr(s, name, None)
        return str(v) if v is not None else None
    return getter


# Definición de filas de la tabla comparativa
COMPARISON_ROWS: list[tuple[str, str, str, callable]] = [
    # (categoría, clave, label, getter)
    # --- Identidad ---
    ("Identidad", "designation", "Designación", lambda s: s.designation_modern),
    ("Identidad", "family", "Familia", lambda s: s.family.family_code),
    ("Identidad", "source", "Fuente", lambda s: s.source_catalog),
    ("Identidad", "available_sack", "Disp. SACK", lambda s: "Sí" if s.available_sack else "—"),
    ("Identidad", "available_cintac", "Disp. CINTAC", lambda s: "Sí" if s.available_cintac else "—"),
    ("Identidad", "is_custom", "Custom", lambda s: "Sí" if s.is_custom else "No"),

    # --- Geometría ---
    ("Geometría", "d", "d (mm)", _get_float("d")),
    ("Geometría", "bf", "bf (mm)", _get_float("bf")),
    ("Geometría", "tf", "tf (mm)", _get_float("tf")),
    ("Geometría", "tw", "tw (mm)", _get_float("tw")),
    ("Geometría", "h", "h (mm)", _get_float("h")),
    ("Geometría", "k", "k (mm)", _get_float("k")),
    ("Geometría", "r", "r (mm)", _get_float("r")),
    ("Geometría", "area", "Área (mm²)", _get_float("area_mm2")),
    ("Geometría", "weight", "Peso (kg/m)", _get_float("weight_kg_m")),
    ("Geometría", "perimeter", "Perímetro (mm)", _get_float("perimeter_mm")),

    # --- Eje fuerte X-X ---
    ("Eje fuerte (X-X)", "Ix", "Ix (mm⁴)", _get_float("Ix_mm4")),
    ("Eje fuerte (X-X)", "Sx", "Sx (mm³)", _get_float("Sx_mm3")),
    ("Eje fuerte (X-X)", "Zx", "Zx (mm³)", _get_float("Zx_mm3")),
    ("Eje fuerte (X-X)", "rx", "rx (mm)", _get_float("rx_mm")),

    # --- Eje débil Y-Y ---
    ("Eje débil (Y-Y)", "Iy", "Iy (mm⁴)", _get_float("Iy_mm4")),
    ("Eje débil (Y-Y)", "Sy", "Sy (mm³)", _get_float("Sy_mm3")),
    ("Eje débil (Y-Y)", "Zy", "Zy (mm³)", _get_float("Zy_mm3")),
    ("Eje débil (Y-Y)", "ry", "ry (mm)", _get_float("ry_mm")),

    # --- Torsión ---
    ("Torsión", "J", "J (mm⁴)", _get_float("J_mm4")),
    ("Torsión", "Cw", "Cw (mm⁶)", _get_float("Cw_mm6")),

    # --- Pandeo local ---
    ("Pandeo local", "bf_2tf", "bf/2tf", _get_float("bf_2tf")),
    ("Pandeo local", "h_tw", "h/tw", _get_float("h_tw")),
    ("Pandeo local", "b_t", "b/t", _get_float("b_t")),
    ("Pandeo local", "D_t", "D/t", _get_float("D_t")),
    ("Pandeo local", "Qs", "Qs", _get_float("Qs")),
    ("Pandeo local", "Qa", "Qa", _get_float("Qa")),
    ("Pandeo local", "Fy", "Fy (MPa)", _get_float("Fy_default_MPa")),
]


@dataclass
class ComparisonCell:
    """Una celda de la tabla comparativa."""
    value: Optional[float | str]
    formatted: str
    diff_pct: Optional[float] = None  # % de diferencia respecto al promedio
    highlight: str = "none"  # "none", "low", "high" (diferencia >5%), "extreme" (>20%)


@dataclass
class Comparison:
    """Resultado de comparar 2-4 perfiles.

    Atributos:
        sections: lista de Section comparados.
        rows: lista de (categoría, label, clave) en orden.
        cells: matriz [n_rows][n_sections] de ComparisonCell.
    """
    sections: list[Section]
    rows: list[tuple[str, str, str]]  # (categoría, label, clave)
    cells: list[list[ComparisonCell]] = field(default_factory=list)

    @classmethod
    def create(cls, sections: list[Section]) -> "Comparison":
        """Crea una comparación desde una lista de 2-4 perfiles."""
        if len(sections) < 2:
            raise ValueError("Se necesitan al menos 2 perfiles para comparar")
        if len(sections) > 4:
            raise ValueError("Máximo 4 perfiles para comparar")

        rows = [(cat, label, key) for cat, key, label, _ in COMPARISON_ROWS]
        cells: list[list[ComparisonCell]] = []

        for cat, key, label, getter in COMPARISON_ROWS:
            row_cells: list[ComparisonCell] = []
            # Obtener valores crudos
            raw_values = []
            for s in sections:
                v = getter(s)
                raw_values.append(v)

            # Calcular promedio para celdas numéricas
            numeric_values = [v for v in raw_values if isinstance(v, (int, float)) and v is not None]
            avg = sum(numeric_values) / len(numeric_values) if numeric_values else None

            for v in raw_values:
                formatted = _format_value(v)
                diff_pct = None
                highlight = "none"

                if isinstance(v, (int, float)) and v is not None and avg and avg != 0:
                    diff_pct = ((v - avg) / abs(avg)) * 100
                    if abs(diff_pct) > 20:
                        highlight = "extreme"
                    elif abs(diff_pct) > 5:
                        highlight = "high"

                row_cells.append(ComparisonCell(
                    value=v,
                    formatted=formatted,
                    diff_pct=diff_pct,
                    highlight=highlight,
                ))
            cells.append(row_cells)

        return cls(sections=sections, rows=rows, cells=cells)

    # ------------------------------------------------------------------
    # Exportación
    # ------------------------------------------------------------------
    def to_dict(self) -> dict:
        """Serializa a diccionario (JSON-ready)."""
        return {
            "sections": [s.designation_modern for s in self.sections],
            "rows": [
                {
                    "categoria": cat,
                    "propiedad": label,
                    "valores": [c.formatted for c in row_cells],
                    "highlights": [c.highlight for c in row_cells],
                }
                for (cat, label, _), row_cells in zip(self.rows, self.cells)
            ],
        }

    def to_tsv(self) -> str:
        """Genera texto TSV (tab-separated) listo para pegar en Excel."""
        # Header: Propiedad + designaciones
        headers = ["Propiedad"] + [s.designation_modern for s in self.sections]
        lines = ["\t".join(headers)]

        current_cat = ""
        for (cat, label, _), row_cells in zip(self.rows, self.cells):
            # Insertar separador de categoría
            if cat != current_cat:
                lines.append(f"--- {cat} ---")
                current_cat = cat
            row = [label] + [c.formatted for c in row_cells]
            lines.append("\t".join(row))

        return "\n".join(lines)

    def to_dataframe(self):
        """Convierte a pandas DataFrame.

        Requiere pandas instalado.
        """
        try:
            import pandas as pd
        except ImportError as exc:
            raise ImportError("pandas es requerido para to_dataframe()") from exc

        data = {}
        for i, s in enumerate(self.sections):
            col_name = s.designation_modern
            data[col_name] = [c.formatted for c in self.cells_as_column(i)]

        # El índice son las propiedades
        index = [label for _, label, _ in self.rows]
        df = pd.DataFrame(data, index=index)
        return df

    def to_excel(self, filepath: str) -> None:
        """Exporta a Excel .xlsx con formato y colores.

        Requiere openpyxl instalado.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("openpyxl es requerido para to_excel()") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparación"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cat_font = Font(bold=True, color="1D1D1F")
        cat_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="E0E0E0"),
        )
        center_align = Alignment(horizontal="right")
        left_align = Alignment(horizontal="left")

        # Header row
        ws.cell(row=1, column=1, value="Propiedad")
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        for i, s in enumerate(self.sections):
            cell = ws.cell(row=1, column=2 + i, value=s.designation_modern)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        current_cat = ""
        row_idx = 2
        for (cat, label, _), row_cells in zip(self.rows, self.cells):
            # Separador de categoría
            if cat != current_cat:
                ws.cell(row=row_idx, column=1, value=f"--- {cat} ---")
                for c in range(1, 2 + len(self.sections)):
                    ws.cell(row=row_idx, column=c).font = cat_font
                    ws.cell(row=row_idx, column=c).fill = cat_fill
                row_idx += 1
                current_cat = cat

            # Propiedad
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = Font(bold=False)
            cell.border = thin_border
            cell.alignment = left_align

            # Valores
            for i, c in enumerate(row_cells):
                vcell = ws.cell(row=row_idx, column=2 + i, value=c.formatted)
                vcell.border = thin_border
                vcell.alignment = center_align
                if c.highlight == "extreme":
                    vcell.fill = orange_fill
                elif c.highlight == "high":
                    vcell.fill = yellow_fill

            row_idx += 1

        # Auto-ajustar columnas
        for col in range(1, 2 + len(self.sections)):
            max_len = 10
            for row in range(1, row_idx):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = max_len + 4

        wb.save(filepath)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def cells_as_column(self, col_idx: int) -> list[ComparisonCell]:
        """Retorna las celdas de una columna (perfil) específica."""
        return [row[col_idx] for row in self.cells]

    @property
    def n_sections(self) -> int:
        return len(self.sections)

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def categories(self) -> list[str]:
        """Lista de categorías únicas en orden."""
        seen = []
        for cat, _, _ in self.rows:
            if cat not in seen:
                seen.append(cat)
        return seen


# ----------------------------------------------------------------------
# Formato de valores
# ----------------------------------------------------------------------
def _format_value(value) -> str:
    """Formatea un valor para mostrar en la tabla comparativa."""
    if value is None:
        return "—"
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        if value == 0:
            return "0"
        if abs(value) >= 1_000_000:
            return f"{value:,.0f}".replace(",", ".")
        if abs(value) >= 1000:
            if value == int(value):
                return f"{int(value):,}".replace(",", ".")
            return f"{value:,.1f}".replace(",", ".")
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    return str(value)
